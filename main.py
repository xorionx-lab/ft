from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, Column, Integer, String, Boolean, DateTime
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from kivy.core.window import Window
from datetime import datetime
from kivy.core.window import Window

Window.size = (Window.width, Window.height)
Window.size = (1920, 1080)

Base = declarative_base()

# Configuração do banco de dados
engine = create_engine("sqlite:///alunos.db", echo=True)
Session = sessionmaker(bind=engine)
session = Session()

# Modelos
class Aluno(Base):
    __tablename__ = "alunos"
    id = Column(Integer, primary_key=True)
    nome = Column(String)
    serie = Column(String)
    turma = Column(String)
    turno = Column(String)
    data_atendimento = Column(DateTime, nullable=True)
    data_reuniao = Column(DateTime, nullable=True)
    demanda = Column(String)
    suporte = Column(String)
    retorno = Column(String)
    horario_atendimento = Column(String)
    resolucao = Column(String)

class Relatorio(Base):
    __tablename__ = "relatorios"
    id = Column(Integer, primary_key=True)
    aluno_id = Column(Integer)
    data_solicitacao = Column(DateTime, nullable=False)
    data_entrega = Column(DateTime, nullable=True)
    profissional_solicitante = Column(String)
    entregue = Column(Boolean, default=False)

# Cria as tabelas no banco de dados
Base.metadata.create_all(engine)

# Função para gerar planilha de alunos
def gerar_planilha():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Alunos"

    # Cabeçalhos da planilha
    headers = ["ID", "Nome", "Série", "Turma", "Turno", "Data Atendimento", "Data Reunião", "Demanda", "Suporte", "Retorno", "Horário Atendimento", "Resolução"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Preencher a planilha com os dados dos alunos
    alunos = session.query(Aluno).all()
    for row_num, aluno in enumerate(alunos, 2):
        sheet[f"A{row_num}"] = aluno.id
        sheet[f"B{row_num}"] = aluno.nome
        sheet[f"C{row_num}"] = aluno.serie
        sheet[f"D{row_num}"] = aluno.turma
        sheet[f"E{row_num}"] = aluno.turno
        sheet[f"F{row_num}"] = aluno.data_atendimento.strftime("%d/%m/%Y") if aluno.data_atendimento else ""
        sheet[f"G{row_num}"] = aluno.data_reuniao.strftime("%d/%m/%Y") if aluno.data_reuniao else ""
        sheet[f"H{row_num}"] = aluno.demanda
        sheet[f"I{row_num}"] = aluno.suporte
        sheet[f"J{row_num}"] = aluno.retorno
        sheet[f"K{row_num}"] = aluno.horario_atendimento
        sheet[f"L{row_num}"] = aluno.resolucao

    # Salvar a planilha
    wb.save("alunos.xlsx")
    print("Planilha gerada com sucesso!")

# Função para adicionar aluno
def adicionar_aluno(nome, serie, turma, turno, data_atendimento, data_reuniao, demanda, suporte, retorno, horario_atendimento, resolucao):
    # Convertendo as datas de string para datetime
    try:
        data_atendimento = datetime.strptime(data_atendimento, "%d/%m/%Y").date() if data_atendimento else None
        data_reuniao = datetime.strptime(data_reuniao, "%d/%m/%Y").date() if data_reuniao else None
    except ValueError:
        print("Formato de data inválido!")
        return

    aluno = Aluno(
        nome=nome,
        serie=serie,
        turma=turma,
        turno=turno,
        data_atendimento=data_atendimento,
        data_reuniao=data_reuniao,
        demanda=demanda,
        suporte=suporte,
        retorno=retorno,
        horario_atendimento=horario_atendimento,
        resolucao=resolucao
    )
    session.add(aluno)
    session.commit()

# Interface do Kivy
class AlunoApp(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical')

        # Campos de entrada
        self.nome_input = TextInput(hint_text="Nome", multiline=False)
        self.serie_input = TextInput(hint_text="Série", multiline=False)
        self.turma_input = TextInput(hint_text="Turma", multiline=False)
        self.turno_input = TextInput(hint_text="Turno", multiline=False)
        self.data_atendimento_input = TextInput(hint_text="Data Atendimento (DD/MM/AAAA)", multiline=False)
        self.data_reuniao_input = TextInput(hint_text="Data Reunião (DD/MM/AAAA)", multiline=False)
        self.demanda_input = TextInput(hint_text="Demanda", multiline=False)
        self.suporte_input = TextInput(hint_text="Suporte", multiline=False)
        self.retorno_input = TextInput(hint_text="Retorno", multiline=False)
        self.horario_atendimento_input = TextInput(hint_text="Horário Atendimento", multiline=False)
        self.resolucao_input = TextInput(hint_text="Resolução", multiline=False)

        # Botões
        self.salvar_button = Button(text="Salvar Aluno", on_press=self.salvar_aluno)
        self.gerar_planilha_button = Button(text="Gerar Planilha", on_press=self.gerar_planilha)

        # Tabela de alunos
        self.tabela_layout = GridLayout(cols=1, size_hint_y=None)
        self.tabela_layout.bind(minimum_height=self.tabela_layout.setter('height'))

        scroll_view = ScrollView(size_hint=(1, None), size=(Window.width, Window.height))
        scroll_view.add_widget(self.tabela_layout)

        # Adicionar widgets ao layout
        self.layout.add_widget(self.nome_input)
        self.layout.add_widget(self.serie_input)
        self.layout.add_widget(self.turma_input)
        self.layout.add_widget(self.turno_input)
        self.layout.add_widget(self.data_atendimento_input)
        self.layout.add_widget(self.data_reuniao_input)
        self.layout.add_widget(self.demanda_input)
        self.layout.add_widget(self.suporte_input)
        self.layout.add_widget(self.retorno_input)
        self.layout.add_widget(self.horario_atendimento_input)
        self.layout.add_widget(self.resolucao_input)
        self.layout.add_widget(self.salvar_button)
        self.layout.add_widget(self.gerar_planilha_button)
        self.layout.add_widget(scroll_view)

        self.carregar_alunos()
        return self.layout

    def salvar_aluno(self, instance):
        nome = self.nome_input.text
        serie = self.serie_input.text
        turma = self.turma_input.text
        turno = self.turno_input.text
        data_atendimento = self.data_atendimento_input.text
        data_reuniao = self.data_reuniao_input.text
        demanda = self.demanda_input.text
        suporte = self.suporte_input.text
        retorno = self.retorno_input.text
        horario_atendimento = self.horario_atendimento_input.text
        resolucao = self.resolucao_input.text

        adicionar_aluno(nome, serie, turma, turno, data_atendimento, data_reuniao, demanda, suporte, retorno, horario_atendimento, resolucao)
        self.carregar_alunos()

    def carregar_alunos(self):
        self.tabela_layout.clear_widgets()
        alunos = session.query(Aluno).all()
        for aluno in alunos:
            row = BoxLayout(size_hint_y=None, height=40)
            row.add_widget(Label(text=str(aluno.id)))
            row.add_widget(Label(text=aluno.nome))
            row.add_widget(Label(text=aluno.serie))
            row.add_widget(Label(text=aluno.turma))
            row.add_widget(Label(text=aluno.turno))
            self.tabela_layout.add_widget(row)

    def gerar_planilha(self, instance):
        gerar_planilha()

if __name__ == '__main__':
    AlunoApp().run()
